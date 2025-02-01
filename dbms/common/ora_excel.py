import os
import pandas as pd
from tkinter import filedialog
from tkinter import messagebox
from tkinter import *
from openpyxl import load_workbook

################################
# 변수 입력 #
model_name = "ABC123"
date = '220901'
################################

root = Tk()
dir_path = filedialog.askdirectory(parent=root, initialdir='C:/', title='폴더를 선택하세요')

if dir_path == '':
    messagebox.showwarning("경고","파일을 추가 하세요")
    
file_list = os.listdir(dir_path)
file_name = ["wb1", "wb2", "wb3"]
for i in range(3):
    print(i,"-",file_name[i],"-",file_list[i])
    globals()[file_name[i]] = load_workbook(dir_path+'/'+file_list[i])



cp_ws = wb1['Sheet1']
cp1_ws = wb2['Sheet1']
py_ws1 = wb3.worksheets[0]
py_ws2 = wb3.worksheets[1]
py_ws3 = wb3.worksheets[2]


py_ws1.title = model_name+'_'+date
py_ws2.title = model_name+'_'+date+'_1'
py_ws3.title = model_name+'_'+date+'_2'



for row in range(2,7):
    for col in range(1,3):
        cell1 = cp_ws.cell(row+4, col+2).value
        cell2 = cp1_ws.cell(row+4,col+2).value
        py_ws2.cell(row,col,value=cell1)
        py_ws3.cell(row,col,value=cell2)


wb3.save(dir_path+'/'+file_list[2])