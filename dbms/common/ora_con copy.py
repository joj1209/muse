import oracledb
import pandas as pd
import datetime
from openpyxl import load_workbook

un = 'US_RCIS_DW'
cs = '10.10.10.11/ORCLCDB'
port = 1521
pw = 'PW_RCIS_DW'

sql1 = """SELECT COUNT(1) AS 건수, SUM(PRDC_CNT) AS 기준값합계 
    FROM (
    SELECT 1 AS PRDC_CNT -- 작품수
    FROM T_DW_BLDN_FIART_PRDC_INFO_N  -- 기준값합계
    WHERE DTY_SPRTN_CD    <> 'D'
    AND LINK_PRCS_ST_CD <> 'F'
    )
"""

sql2 = """SELECT COUNT(1) AS 건수, SUM(STRY_CNTNS_DATA_CNT) AS 기준값합계 
    FROM (
    SELECT 1 AS STRY_CNTNS_DATA_CNT -- 이야기컨텐츠자료수
    FROM T_DW_CLTR_PLNG_CNTNS_INFO_N  -- 문화기획컨텐츠정보
    )
"""

sql1 = [sql1, sql2]


with oracledb.connect(user=un, password=pw, dsn=cs, port=port) as connection:
    with connection.cursor() as cursor:
        sql = """select to_char(sysdate, 'yyyymmdd hh24:mi:ss'), instance_name, version from v$instance"""
        for r in cursor.execute(sql):
            print(r)
    # connection.close()
    

# DB Connection Info
db_connections = [
    {'un': 'US_RCIS_DW', 'pw': 'PW_RCIS_DW', 'cs': '10.10.10.11/ORCLCDB', 'port': 1521}
    # {'un': 'system', 'pw': 'oracle', 'cs': '192.168.137.19/oracle19', 'port': 1521}
]

# Set Output Filename
current_time = datetime.datetime.now()
formatted_time = current_time.strftime("%Y%m%d_%H%M")
output_filename = f'output_{formatted_time}.xlsx'

# Extract Data and Create DataFrame
df_list = []
for db_info in db_connections:
    un = db_info['un']
    pw = db_info['pw']
    cs = db_info['cs']
    port = db_info['port']
    
    for sql in sql1:
        print(sql)

        with oracledb.connect(user=un, password=pw, dsn=cs, port=port) as connection:
            with connection.cursor() as cursor:
                # sql = "SELECT * FROM v$instance"
                cursor.execute(sql)

                columns = [i[0] for i in cursor.description]
                rows = cursor.fetchall()

                df = pd.DataFrame(rows, columns=columns)
                
        df_list.append(df)
        
# Combine DataFrames if there are multiple
if len(df_list) > 1:
    combined_df = pd.concat(df_list, ignore_index=True)
else:
    combined_df = df_list[0]

# Write DataFrame to Excel
# combined_df.to_excel(output_filename, index=False)

print(f"Excel file '{output_filename}' has been created successfully.")

# 엑셀 파일 불러오기
wb = load_workbook(filename='단위테스트결과서.xlsx')

# 첫 번째 시트 불러오기
ws = wb['단위테스트결과서(DM)']

i=0
for row in ws.rows:
    # print(row)
    i=i+1
    l=list(map(int,str(i)))
    
    
print(l)

for j in range(5,12):
    print(j)
    
    
for db_info in db_connections:
    un = db_info['un']
    pw = db_info['pw']
    cs = db_info['cs']
    port = db_info['port']
    
    for j in range(5,12):
        row_num = "AH"+"{}".format(j)
        print(row_num)
        sql = ws[row_num].value
        print(sql)

        with oracledb.connect(user=un, password=pw, dsn=cs, port=port) as connection:
            with connection.cursor() as cursor:
                # sql = "SELECT * FROM v$instance"
                cursor.execute(sql)

                columns = [i[0] for i in cursor.description]
                rows = cursor.fetchall()

                df = pd.DataFrame(rows, columns=columns)
                
    df_list.append(df)    

combined_df.to_excel(output_filename, index=False)

print(ws.max_row)

# print(ws)
excel_sql=ws['AH1'].value

seq = [1-3]

# with oracledb.connect(user=un, password=pw, dsn=cs, port=port) as connection:
#     with connection.cursor() as cursor:
#         sql = """select to_char(sysdate, 'yyyymmdd hh24:mi:ss'), instance_name, version from v$instance"""
#         for r in cursor.execute(excel_sql):
#             print(r)
    # connection.close()
