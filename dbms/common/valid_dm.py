import oracledb
import pandas as pd
import datetime
from openpyxl import load_workbook

un = 'US_RCIS_DW'
cs = '10.10.10.11/ORCLCDB'
port = 1521
pw = 'PW_RCIS_DW'

# DB Connection Info
db_connections = [
    {'un': 'US_RCIS_DW', 'pw': 'PW_RCIS_DW', 'cs': '10.10.10.11/ORCLCDB', 'port': 1521}
    # {'un': 'system', 'pw': 'oracle', 'cs': '192.168.137.19/oracle19', 'port': 1521}
]

# Set Output Filename
current_time = datetime.datetime.now()
formatted_time = current_time.strftime("%Y%m%d_%H%M")
output_filename = f'output_{formatted_time}.xlsx'

# Excel load
wb = load_workbook(filename='단위테스트결과서_working1.xlsx')

# '단위테스트결과서(DM)' sheet load
ws = wb['단위테스트결과서(DM)']

# Extract Data and Create DataFrame
df_list = []    
for db_info in db_connections:
    un = db_info['un']
    pw = db_info['pw']
    cs = db_info['cs']
    port = db_info['port']
    
    for j in range(5,33):
    # for j in range(33,67):    --5,67
        row_num_s = "AH"+"{}".format(j)
        row_num_t = "AI"+"{}".format(j)
        case_id = "B"+"{}".format(j)
        # print(row_num)
        sql_s = ws[row_num_s].value
        sql_t = ws[row_num_t].value
        case_id = ws[case_id].value
        # print(sql)
        # print(case_id)

        with oracledb.connect(user=un, password=pw, dsn=cs, port=port) as connection:
            with connection.cursor() as cursor:

                # 소스sql    
                cursor.execute(sql_s)                
                columns_x = [i[0] for i in cursor.description]
                
                # 마트는 검증컬럼2개,코드는 검증컬럼1개
                if len(columns_x) == 1:
                    columns = ['소스_건수']
                else:
                    columns = ['소스_건수','소스_합계']         
                    
                rows = cursor.fetchall()
                df_s = pd.DataFrame(rows, columns=columns)
                df_s['단위테스트ID'] = case_id       
                print(df_s)         
                
                # 타겟sql
                cursor.execute(sql_t)
                
                if len(cursor.description) == 1:
                    columns = ['타겟_건수']
                else:
                    columns = ['타겟_건수','타겟_합계']
            
                rows = cursor.fetchall()
                df_t = pd.DataFrame(rows, columns=columns)
                df_t['단위테스트ID'] = case_id
                
                # 소스+타겟 결과 합치기
                df = pd.merge(left=df_t, right=df_s, how='inner', on='단위테스트ID') 
                df['타겟_합계'] = df['타겟_합계'].fillna(0)
                
                # 컬럼순서조정
                if len(columns_x) == 1:
                    df = df[['단위테스트ID','소스_건수','타겟_건수']]
                    df['소스_합계'] = 0
                    df['타겟_합계'] = 0
                else:
                    df = df[['단위테스트ID','소스_건수','타겟_건수','소스_합계','타겟_합계']]
                
                sheet = wb._sheets[0]
                sheet["AU5"].value = "aaa"
                
                
        df_list.append(df)
        print(df_list)    

# Combine DataFrames if there are multiple
if len(df_list) > 1:
    combined_df = pd.concat(df_list, ignore_index=True)
else:
    combined_df = df_list[0]


# combined_df.to_excel(output_filename, index=False)

print(ws.max_row)

# print(ws)
excel_sql=ws['AH1'].value

# seq = [1-3]
