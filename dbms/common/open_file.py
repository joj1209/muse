import tkinter as tk
import tkinter.filedialog #이걸 따로 추가해야된다. 왜인지 모르겠음;
from openpyxl import load_workbook

#윈도우 생성
root=tk.Tk()

#전체 이름
root.title('open file')

#창 크기 +붙은 부분은 좌상단 떨어진 위치
root.geometry("1000x400+100+100")

filename=0

def openfile():
    global filename
    filename=tk.filedialog.askopenfilename(initialdir = "C:\MyWork\03. Project\06. 시험",
        title = "open file", filetypes = (("excel file","*.xlsx"),("all files","*.*")))

lab00=tk.Label(root,text="파일 불러오기",font=('Arial 20 bold'),bg='white',fg="black",width=11,height=1)
lab00.grid(row=0,column=0,padx=5,pady=20)

#이미지 추가
file_img=tk.PhotoImage(file="C:/MyWork/cap_p.png")
file_img=file_img.subsample(2,2)


button01 = tkinter.Button(root, overrelief="solid", command=openfile,bg="white",image=file_img)
button01.grid(row=0,column=1,padx=5,pady=20)


while True : 
    if filename != 0:
        break
    root.update() 


lab02=tk.Label(root,text=filename,font=('Arial 12 bold'),bg='white',fg="black",width=50,height=1)
lab02.grid(row=0,column=2,padx=5,pady=20)


# Excel load
# wb = load_workbook(filename='단위테스트결과서_working.xlsx', data_only=True)/
wb = load_workbook(filename=filename, data_only=True)

# '단위테스트결과서(DM)' sheet load
ws = wb['단위테스트결과서(DW)']

root.mainloop()